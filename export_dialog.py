"""Export dialog for exporting calendar events to Excel."""

import os
import sys
from datetime import datetime, timedelta
from PySide6.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QFormLayout,
    QDateTimeEdit,
    QPushButton,
    QHBoxLayout,
    QLabel,
    QFileDialog,
    QMessageBox,
    QRadioButton,
    QButtonGroup,
    QFrame,
)
from PySide6.QtCore import Qt, Signal

from event_card import parse_event_time, is_all_day_event


def get_desktop_path() -> str:
    """Get the user's desktop path."""
    return os.path.join(os.path.expanduser("~"), "Desktop")


def export_events_to_excel(events: list, file_path: str) -> bool:
    """Export events to an Excel file using openpyxl.

    Returns True on success, False on failure.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        # Fallback: export to CSV if openpyxl not available
        return export_events_to_csv(events, file_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "飞书日程"

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = "飞书日程导出"
    ws["A1"].font = Font(size=14, bold=True, color="4B3FE3")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    headers = ["序号", "日期", "开始时间", "结束时间", "时长", "标题", "组织者", "会议链接"]
    # Expand to 8 columns
    ws.merge_cells("A1:H1")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B3FE3", end_color="4B3FE3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[2].height = 22

    # Data rows
    for idx, ev in enumerate(events, 1):
        row = idx + 2  # start from row 3

        start = parse_event_time(ev.get("start_time", {}))
        end = parse_event_time(ev.get("end_time", {}))
        all_day = is_all_day_event(ev)
        summary = str(ev.get("summary", "(无标题)"))

        organizer = ev.get("event_organizer", {})
        organizer_name = ""
        if isinstance(organizer, dict):
            organizer_name = str(organizer.get("display_name", ""))

        vchat = ev.get("vchat", {})
        meeting_url = ""
        if isinstance(vchat, dict):
            meeting_url = str(vchat.get("meeting_url", ""))

        # Duration
        if all_day:
            duration_str = "全天"
        else:
            duration = end - start
            hours = int(duration.total_seconds() // 3600)
            minutes = int((duration.total_seconds() % 3600) // 60)
            duration_str = ""
            if hours > 0:
                duration_str += f"{hours}小时"
            if minutes > 0:
                duration_str += f"{minutes}分钟"

        # Values
        date_str = start.strftime("%Y-%m-%d")
        start_str = "全天" if all_day else start.strftime("%H:%M")
        end_str = "全天" if all_day else end.strftime("%H:%M")

        values = [idx, date_str, start_str, end_str, duration_str, summary, organizer_name, meeting_url]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))

    # Auto-adjust column widths
    col_widths = [6, 14, 12, 12, 12, 30, 14, 40]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A2:H{len(events) + 2}"

    # Freeze header
    ws.freeze_panes = "A3"

    try:
        wb.save(file_path)
        return True
    except PermissionError:
        return False


def export_events_to_csv(events: list, file_path: str) -> bool:
    """Fallback: export events to CSV."""
    import csv
    # Change extension to .csv
    csv_path = file_path.rsplit(".", 1)[0] + ".csv"
    try:
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["序号", "日期", "开始时间", "结束时间", "时长", "标题", "组织者", "会议链接"])
            for idx, ev in enumerate(events, 1):
                start = parse_event_time(ev.get("start_time", {}))
                end = parse_event_time(ev.get("end_time", {}))
                all_day = is_all_day_event(ev)
                summary = str(ev.get("summary", "(无标题)"))
                organizer = ev.get("event_organizer", {})
                organizer_name = str(organizer.get("display_name", "")) if isinstance(organizer, dict) else ""
                vchat = ev.get("vchat", {})
                meeting_url = str(vchat.get("meeting_url", "")) if isinstance(vchat, dict) else ""

                if all_day:
                    duration_str = "全天"
                    start_str = "全天"
                    end_str = "全天"
                else:
                    duration = end - start
                    hours = int(duration.total_seconds() // 3600)
                    minutes = int((duration.total_seconds() % 3600) // 60)
                    duration_str = f"{hours}小时{minutes}分钟" if hours or minutes else ""
                    start_str = start.strftime("%H:%M")
                    end_str = end.strftime("%H:%M")

                writer.writerow([
                    idx, start.strftime("%Y-%m-%d"), start_str, end_str,
                    duration_str, summary, organizer_name, meeting_url,
                ])
        return True
    except Exception:
        return False


class ExportDialog(QDialog):
    """Dialog for exporting events to Excel."""

    def __init__(self, events: list, current_date: datetime, parent=None):
        super().__init__(parent)
        self._events = events
        self._current_date = current_date
        self.setWindowTitle("导出日程到 Excel")
        self.setFixedSize(440, 420)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        title = QLabel("导出日程")
        title.setObjectName("detailTitle")
        layout.addWidget(title)

        # Date range
        range_group = QFrame()
        range_group.setObjectName("dayCell")
        range_layout = QVBoxLayout(range_group)
        range_layout.setContentsMargins(12, 8, 12, 8)
        range_layout.setSpacing(8)

        range_label = QLabel("选择导出范围：")
        range_label.setObjectName("detailLabel")
        range_layout.addWidget(range_label)

        self.range_group = QButtonGroup(self)

        self.radio_month = QRadioButton("当月（{0}）".format(
            self._current_date.strftime("%Y年%m月")
        ))
        self.radio_month.setChecked(True)
        self.range_group.addButton(self.radio_month, 0)
        range_layout.addWidget(self.radio_month)

        self.radio_week = QRadioButton("本周")
        self.range_group.addButton(self.radio_week, 1)
        range_layout.addWidget(self.radio_week)

        self.radio_custom = QRadioButton("自定义范围")
        self.range_group.addButton(self.radio_custom, 2)
        range_layout.addWidget(self.radio_custom)

        # Custom date range inputs
        custom_layout = QFormLayout()
        custom_layout.setContentsMargins(20, 0, 0, 0)
        custom_layout.setSpacing(6)

        self.start_input = QDateTimeEdit()
        self.start_input.setDisplayFormat("yyyy-MM-dd")
        self.start_input.setCalendarPopup(True)
        self.start_input.setDateTime(self._current_date)
        self.start_input.setEnabled(False)
        custom_layout.addRow("从", self.start_input)

        self.end_input = QDateTimeEdit()
        self.end_input.setDisplayFormat("yyyy-MM-dd")
        self.end_input.setCalendarPopup(True)
        self.end_input.setDateTime(self._current_date + timedelta(days=7))
        self.end_input.setEnabled(False)
        custom_layout.addRow("到", self.end_input)

        range_layout.addLayout(custom_layout)

        self.radio_custom.toggled.connect(self._on_custom_toggled)
        self.radio_month.toggled.connect(self._on_range_changed)
        self.radio_week.toggled.connect(self._on_range_changed)

        layout.addWidget(range_group)

        # File path
        path_layout = QHBoxLayout()
        path_label = QLabel("保存位置：")
        path_label.setObjectName("detailLabel")
        path_layout.addWidget(path_label)

        self.path_input = QLabel()
        self.path_input.setObjectName("detailValue")
        self.path_input.setWordWrap(True)
        default_name = f"飞书日程_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self._default_path = os.path.join(get_desktop_path(), default_name)
        self.path_input.setText(self._default_path)
        path_layout.addWidget(self.path_input, 1)

        browse_btn = QPushButton("浏览...")
        browse_btn.setObjectName("secondaryBtn")
        browse_btn.clicked.connect(self._on_browse)
        path_layout.addWidget(browse_btn)
        layout.addLayout(path_layout)

        # Event count preview
        self.count_label = QLabel()
        self.count_label.setObjectName("detailLabel")
        self._on_range_changed()
        layout.addWidget(self.count_label)

        layout.addStretch()

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.setObjectName("secondaryBtn")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)

        export_btn = QPushButton("导出")
        export_btn.setObjectName("primaryBtn")
        export_btn.clicked.connect(self._on_export)
        btn_row.addWidget(export_btn)

        layout.addLayout(btn_row)

    def _on_custom_toggled(self, checked):
        self.start_input.setEnabled(checked)
        self.end_input.setEnabled(checked)
        self._on_range_changed()

    def _on_range_changed(self):
        events = self._get_filtered_events()
        self.count_label.setText(f"将导出 {len(events)} 条日程")

    def _get_date_range(self) -> tuple:
        """Return (start, end) datetime for the selected range."""
        btn_id = self.range_group.checkedId()
        if btn_id == 0:  # Current month
            start = self._current_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if self._current_date.month == 12:
                end = start.replace(year=start.year + 1, month=1) - timedelta(seconds=1)
            else:
                end = start.replace(month=start.month + 1) - timedelta(seconds=1)
        elif btn_id == 1:  # Current week (Mon-Sun)
            today = datetime.now()
            monday = today - timedelta(days=today.weekday())
            start = monday.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=6, hours=23, minutes=59, seconds=59)
        else:  # Custom
            start = self.start_input.dateTime().toPython()
            start = start.replace(hour=0, minute=0, second=0, microsecond=0)
            end = self.end_input.dateTime().toPython()
            end = end.replace(hour=23, minute=59, second=59, microsecond=0)
        return start, end

    def _get_filtered_events(self) -> list:
        """Filter events by the selected date range."""
        start, end = self._get_date_range()
        filtered = []
        for ev in self._events:
            ev_start = parse_event_time(ev.get("start_time", {}))
            if start <= ev_start <= end:
                filtered.append(ev)
        return filtered

    def _on_browse(self):
        default_name = f"飞书日程_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        default_path = os.path.join(get_desktop_path(), default_name)
        file_path, _ = QFileDialog.getSaveFileName(
            self, "选择保存位置", default_path,
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv);;所有文件 (*.*)",
        )
        if file_path:
            self._default_path = file_path
            self.path_input.setText(file_path)

    def _on_export(self):
        events = self._get_filtered_events()
        if not events:
            QMessageBox.information(self, "提示", "所选范围内没有日程可导出")
            return

        file_path = self.path_input.text()
        if not file_path:
            QMessageBox.warning(self, "提示", "请选择保存位置")
            return

        # Ensure .xlsx extension
        if not file_path.endswith((".xlsx", ".csv")):
            file_path += ".xlsx"

        success = export_events_to_excel(events, file_path)
        if success:
            QMessageBox.information(
                self, "导出成功",
                f"已导出 {len(events)} 条日程到：\n{file_path}",
            )
            self.accept()
        else:
            QMessageBox.critical(
                self, "导出失败",
                f"无法写入文件：{file_path}\n请检查文件是否被其他程序占用。",
            )
